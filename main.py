import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from datetime import datetime
from colorama import Fore, Style

# Singleton para el WebDriver
class WebDriverSingleton:
    _instance = None

    @staticmethod
    def get_instance():
        if WebDriverSingleton._instance is None:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--disable-gpu")  # Desactivar errores de GPU
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-webgl")
            chrome_options.add_argument("--disable-software-rasterizer")
            chrome_options.add_argument("--disable-accelerated-2d-canvas")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
            
            # Agregar headers
            chrome_options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
            )
            service = Service(ChromeDriverManager().install())
            WebDriverSingleton._instance = webdriver.Chrome(service=service, options=chrome_options)
        return WebDriverSingleton._instance

# Extraer información de productos (reutilizable para diferentes plataformas)
def extract_product_info(container, platform):
    try:
        if platform == "eBay":
            title = container.find_element(By.CSS_SELECTOR, ".s-item__title").text.strip()
            price = container.find_element(By.CSS_SELECTOR, ".s-item__price").text.strip()
            link = container.find_element(By.CSS_SELECTOR, ".s-item__link").get_attribute('href')
        elif platform == "Amazon":
            try:
                title = container.find_element(By.CSS_SELECTOR, "h2 span").text.strip()
            except NoSuchElementException:
                title = container.find_element(By.CSS_SELECTOR, ".a-text-normal").text.strip()  # Ajuste en Amazon
            try:
                price = container.find_element(By.CSS_SELECTOR, ".a-price .a-offscreen").text.strip()
            except NoSuchElementException:  # Manejo de excepción si no se encuentra el precio
                price = container.find_element(By.CSS_SELECTOR, ".a-price .a-price-whole").text.strip()  # Intento alternativo
            link = container.find_element(By.CSS_SELECTOR, "h2 a").get_attribute('href')
        
        price_value = None
        if price:
            try:
                price_value = float(price.replace('$', '').replace(',', '').strip())
            except ValueError:
                price_value = None

        return {
            'title': title,
            'price': price_value,
            'url': link,
            'platform': platform
        }
    except (NoSuchElementException, ValueError) as e:
        print(Fore.YELLOW + f"[Advertencia] Error al extraer datos de {platform}: {str(e)}" + Style.RESET_ALL)
        return None

# Scraping en eBay
def scrape_ebay_with_selenium(product):
    url = f"https://www.ebay.com/sch/i.html?_nkw={product}"
    print(Fore.GREEN + f"\nScraping de eBay para: {product}..." + Style.RESET_ALL)
    driver = WebDriverSingleton.get_instance()
    driver.get(url)

    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".s-item__info.clearfix"))
        )
    except TimeoutException:
        print(Fore.RED + "Tiempo agotado, no se encontraron resultados en eBay." + Style.RESET_ALL)
        return []

    containers = driver.find_elements(By.CSS_SELECTOR, ".s-item__info.clearfix")
    results = [info for c in containers if (info := extract_product_info(c, "eBay"))]
    print(Fore.CYAN + f"eBay: {len(results)} productos encontrados\n" + Style.RESET_ALL)
    return results

# Scraping en Amazon
def scrape_amazon_with_selenium(product):
    url = f"https://www.amazon.com/s?k={product}"
    print(Fore.GREEN + f"Scraping de Amazon para: {product}..." + Style.RESET_ALL)
    driver = WebDriverSingleton.get_instance()
    driver.get(url)

    try:
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".s-main-slot"))
        )
    except TimeoutException:
        print(Fore.RED + "Tiempo agotado, no se encontraron resultados en Amazon." + Style.RESET_ALL)
        return []

    containers = driver.find_elements(By.CSS_SELECTOR, ".s-main-slot .s-result-item")
    results = [info for c in containers if (info := extract_product_info(c, "Amazon"))]
    print(Fore.CYAN + f"Amazon: {len(results)} productos encontrados\n" + Style.RESET_ALL)
    return results

def analyze_prices(data):
    """Analiza los precios y calcula indicadores clave."""
    valid_prices = [item['price'] for item in data if item['price'] is not None]
    if not valid_prices:
        return {}

    analysis = {
        "Precio Promedio": sum(valid_prices) / len(valid_prices),
        "Precio Mínimo": min(valid_prices),
        "Precio Máximo": max(valid_prices),
        "Desviación Estándar": pd.Series(valid_prices).std(),
        "Productos en Amazon": sum(1 for item in data if item['platform'] == "Amazon"),
        "Productos en eBay": sum(1 for item in data if item['platform'] == "eBay"),
    }
    return analysis

# Guardar resultados en Excel
def save_to_excel_with_analysis(data, product, folder="scraping_results"):
    """Guarda los resultados del scraping y un análisis de precios en un archivo Excel."""
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, f"scraping_{product}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = Workbook()

    # Resultados
    ws_results = wb.active
    ws_results.title = "Resultados"
    headers = ['Título', 'Precio', 'Plataforma', 'URL']
    ws_results.append(headers)
    for item in data:
        ws_results.append([item['title'], f"${item['price']:.2f}" if item['price'] else "N/A", item['platform'], item['url']])

    # Estilo de encabezados
    fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws_results[1]:
        cell.fill = fill
        cell.font = font

    # Ajustar ancho
    for col in ws_results.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_results.column_dimensions[col[0].column_letter].width = max_len + 2

    # Análisis de precios
    ws_analysis = wb.create_sheet("Análisis de Precios")
    analysis = analyze_prices(data)
    if analysis:
        ws_analysis.append(["Indicador", "Valor"])
        for key, value in analysis.items():
            if isinstance(value, float):
                value = f"${value:.2f}" if "Precio" in key else f"{value:.2f}"
            ws_analysis.append([key, value])

        # Estilo de encabezados
        for cell in ws_analysis[1]:
            cell.fill = fill
            cell.font = font

        # Ajustar ancho
        for col in ws_analysis.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_analysis.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(file_path)
    print(Fore.YELLOW + f"Archivo guardado en: {file_path}\n" + Style.RESET_ALL)

# Función principal actualizada
def main():
    product = input("\nIntroduce el producto a buscar: ")
    ebay_results = scrape_ebay_with_selenium(product)
    amazon_results = scrape_amazon_with_selenium(product)

    all_results = ebay_results + amazon_results
    if all_results:
        save_to_excel_with_analysis(all_results, product)

    # Cerrar el navegador
    WebDriverSingleton.get_instance().quit()

if __name__ == "__main__":
    main()
